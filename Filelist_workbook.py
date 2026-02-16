# ----------------------------------------
# Name    : Audrey
# Note    : This is for creating a xlsx file for listing a bunch of files in a single folder by last write time.
#
# Date Created : July 15, 2025 @8:50PM
# ----------------------------------------

# public library imports
import datetime
import time
import os.path
import sys
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_points
from PIL import Image as PILImage

class Excel_Filelist:
    """
    A class to handle file operations and Excel spreadsheet management.
    """
    def __init__(self,
                 worksheet=None,
                 dir_path='',
                 screenshots_folder=r'C:\Users\Audrey\OneDrive\Pictures\screenshot-resized100'):
        self.ws = worksheet
        self.dictionary = {}
        self.dir_path = dir_path
        self.screenshots_folder = screenshots_folder

    def get_file_creation_date(self, file_name):
        """
        Returns the file creation date as a human-readable string.
        On Windows: actual creation time.
        On Unix: metadata change time (may not be true creation time).
        """
        file_path = os.path.join(self.dir_path, file_name)

        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")

            # # Get creation time (ctime)
            # creation_time = os.path.getctime(file_path)

            # Get modified time (mtime)
            modification_time = os.path.getmtime(file_path)

            # Convert to human-readable format
            readable_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(modification_time))
            return readable_time

        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            return None

    def import_dictionary(self):
        try:
            all_files = os.listdir(self.dir_path)
            filekey = [f for f in all_files if os.path.isfile(self.dir_path + '/' + f)]

            for f in filekey:  # [:] creates a copy to iterate over
                if f.endswith('.xlsx'):
                    filekey.remove(f)
            timestamps = [self.get_file_creation_date(t) for t in all_files]
            unsorted_dict = dict(zip(filekey, timestamps))
            # sort by value (ascending). value is timestamps
            self.dictionary = dict(sorted(unsorted_dict.items(), key=lambda item: item[1]))

        except:
            return {0}
        return self.dictionary

    def set_column_width_pixels(self, col_letter: str, width: float = 8.43) -> float:
        """
        Set approximate column width on the active worksheet.
        """
        self.ws.column_dimensions[col_letter].width = 2 * width
        return self.ws.column_dimensions[col_letter].width


    def print_dictionary(self):
        """ max_width formats all printed out statements to be a neat spaced out columns.
                item1       item2
                item0000    item2
                ---         ---
        """
        max_width = max( (len(str(f)) for f in self.dictionary.keys()), default=0)
        print(*[f"{f:<{max_width}}\t{t}" for f, t in self.dictionary.items()], sep="\n")

    # image = 2025-02-07 144957 switch_before_go.png
    # image_dir = C:\Users\Audrey\OneDrive\Pictures\screenshot-collages
    # saved_image_dir = C:\Users\Audrey\OneDrive\Pictures\screenshot-resized100
    def filelist_thumbnail(self, idx=0, image=None,
                           image_dir=r"",
                           saved_image_dir=r""):

        # checks extension name if it is .png, .jpg, or .jpeg
        name, extension = os.path.splitext(image)
        if extension.lower() not in ['.png', '.jpg', '.jpeg']:
            #breakpoint()
            return None

        else:
            # Add image to my filelist spreadsheet
            image_path = os.path.join(image_dir, image)
            cell_address = f"A{idx + 3}"

            # Load with Pillow first to get pixel dimensions
            pil_img = PILImage.open(image_path)
            width_px, height_px = pil_img.size
            # print(f"Original Pixel dimensions: {width_px} x {height_px}")

            # Convert Excel column width → pixels (approximation)
            scale = 100
            # Scale height proportionally, height/width
            proportional_factor = height_px / width_px
            target_height_px = int(proportional_factor * scale)
            # print(f"New Pixel dimensions: {scale} x {target_height_px}")

            # (width, height) makes image proportional to uniform width for xlsx sheet.
            pil_img = pil_img.resize((scale, target_height_px))
            saved_image_path = os.path.join(saved_image_dir, image)
            pil_img.save(saved_image_path)
            pil_img.close()  # IMPORTANT: Close PIL image before openpyxl reads it

            # Load into openpyxl and anchor
            img = Image(saved_image_path)

            img.anchor = cell_address
            self.ws.row_dimensions[idx + 3].height = pixels_to_points(target_height_px)
            self.ws.add_image(img)
            # filelist_wb.save(self.ws)
            return saved_image_path

# def add_to_spreadsheet(dictionary, letter1, letter2):
#     return_str = ""
#     i = 3
#     for key, value in dictionary.items():
#         name_cell1 = f"{letter1}{i}"
#         name_cell2 = f"{letter2}{i}"
#         ws[name_cell1] = f"{key}"
#         ws[name_cell2] = f"{value}"
#         return_str = return_str + f"{name_cell1}\t{key}\t\t\t{name_cell2}\t{value}\n"
#         i += 1
#     return return_str
#
# if __name__ == '__main__':
#
#     """
#     Run this in terminal to run args.py:
#
#     (venv) PS {filepath} >      wsl
#     root@DESKTOP-8G4J2N7:/mnt/c/Users/Audrey/.../aluan_file-organizer-py#     python3 Filelist_workbook.py "/mnt/c/Users/Audrey/OneDrive/Pictures/random_images"
#
#
#     """
#
#     print(sys.executable)
#     print(sys.path)
#
#     dt = datetime.datetime.today()
#     date = dt.strftime("%Y %B %d @%I:%M%p")
#
#     screenshots_folder = r"C:\Users\Audrey\OneDrive\Pictures\Clip Studio Paint files"
#
#     # Load existing file
#     filelist_wb = Workbook()
#     ws = filelist_wb.active
#     ws.title = "List of Files"
#
#     ws['B1'] = "Filelist of folder"
#     ws['C1'] = date
#     print(date)
#     #"2025-07-16 11:04AM"
#     ws['A2'] = "Image"
#     ws['B2'] = "Name"
#     ws['C2'] = "LastWriteTime"
#
#     # Pass the worksheet object
#     excel_fl = Excel_Filelist(worksheet=ws, dir_path=screenshots_folder)
#
#     width = 10
#     excel_fl.set_column_width_pixels('A', width)
#     excel_fl.set_column_width_pixels('B', width*2)
#     excel_fl.set_column_width_pixels('C', width*1.5)
#
#     i = 3
#     dictionary = excel_fl.import_dictionary()
#     excel_fl.print_dictionary(dictionary)
#     imgname1, timestamp2 = next(iter(dictionary.items()))
#
#     name_cell = f"B{i}"
#     lastwritetime_cell = f"C{i}"
#
#     ws[name_cell] = f"{imgname1}"
#     ws[lastwritetime_cell] = f"{timestamp2}"
#
#     print("--------------------\n")
#
#     # image_dir = sys.argv[1]
#     image_dir = screenshots_folder
#     saved_image_dir = r"C:\Users\Audrey\OneDrive\Pictures\screenshot-resized100"
#
#     # i is the counter for number of dictionary items. The (image1, timestamp1) unpacks the dictionary items.
#     for i, (image1, timestamp1) in enumerate(dictionary.items()):
#         """
#         uncomment this block of code if you want to debug or update the code, and only test with the first 10 images.
#         """
#         # if i >= 10:
#         #     break
#         """
#         the os path join function is implemented so that every image listed in the dictionary automatically forms each filepath to each individual image
#         in order to embed each thumbnail image into each row of the spreadsheet.
#         """
#         returned_path = excel_fl.filelist_thumbnail(idx=i, image=image1, image_dir=image_dir, saved_image_dir=saved_image_dir)
#         # print(returned_path)
#
#     add_to_spreadsheet(dictionary, "B", "C")
#
#     filelist_wb.save('filelist.xlsx')
#     print(f"\n{ws.title} spreadsheet is saved!")