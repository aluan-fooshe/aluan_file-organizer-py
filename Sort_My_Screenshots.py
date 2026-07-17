# ----------------------------------------
# Name    : Audrey
# Note    : This makes a spreadsheet of all the screenshots in my Screenshots folder, before I turn them into a Canva
#           collage and trash the originals. It will show the amount of GB I removed and the date it was created.
#           Potentially stuff I can add to my future collage projects.
#
# Date Created : August 27, 2025 @2:01PM
# ----------------------------------------

# public library imports
import datetime
import sys
from openpyxl import Workbook
from Filelist_workbook import Excel_Filelist
import os
import re

def add_dict_to_spreadsheet(dictionary, letter1, letter2):
    return_str = ""
    i = 3
    for key, value in dictionary.items():
        name_cell1 = f"{letter1}{i}"
        name_cell2 = f"{letter2}{i}"
        ws[name_cell1] = f"{key}"
        ws[name_cell2] = f"{value}"
        return_str = return_str + f"{name_cell1}\t{key}\t\t\t{name_cell2}\t{value}\n"
        i += 1
    return return_str

def add_value_to_spreadsheet(dictionary, letter1):
    return_str = ""
    i = 3
    for key, value in dictionary.items():
        name_cell1 = f"{letter1}{i}"
        ws[name_cell1] = f"{value}"
        return_str = return_str + f"{name_cell1}\t{value}\n"
        i += 1
    return return_str

if __name__ == "__main__":
    print(sys.executable)
    print(sys.path)

    dt = datetime.datetime.today()
    date = dt.strftime("%Y %B %d @%I:%M%p")

    dir_path = r"C:\Users\Audrey\OneDrive\Pictures\Camera Roll\takeout-20250725T035235Z-1-001\Unsorted_2017-2025\2024_Screenshots"

    # logic for automatically naming the screenshots file list to directory name.
    last_part = os.path.basename(dir_path)
    xlsx_name = last_part.replace(" ", "-")

    # Load existing file
    filelist_wb = Workbook()
    ws = filelist_wb.active
    ws.title = "List of Files"

    ws['B1'] = "Filelist of folder"
    ws['C1'] = dir_path.split("\\OneDrive\\")[1]
    ws['D1'] = date
    print(date)
    # "2025-07-16 11:04AM"
    ws['A2'] = "Image"
    ws['B2'] = "Name"
    ws['C2'] = "LastWriteTime"

    # Pass the worksheet object
    excel_fl = Excel_Filelist(worksheet=ws, dir_path=dir_path)

    width = 10
    C_width = len(str(ws['C1'].value))

    excel_fl.set_column_width_pixels('A', width)
    excel_fl.set_column_width_pixels('B', width * 2)

    # # Added column for renaming all video files
    excel_fl.set_column_width_pixels('C', C_width / 2)
    # excel_fl.set_column_width_pixels('D', width * 1.5)

    i = 3
    excel_fl.import_dictionary()
    imgname1, timestamp2 = next(iter(excel_fl.dictionary.items()))

    name_cell = f"B{i}"
    # new_name_cell = f"C{i}"
    # lastwritetime_cell = f"D{i}"
    lastwritetime_cell = f"C{i}"

    ws[name_cell] = f"{imgname1}"
    ws[lastwritetime_cell] = f"{timestamp2}"

    print("--------------------\n")

    new_names = {}
    new_filename = ""
    for i, (image1, timestamp1) in enumerate(excel_fl.dictionary.items()):
        if image1.endswith(").mp4"):
            print(image1)
            """ remove the parenthesis from filename
                    filename (1).mp4 --> filename1.mp4 """
            new_filename = re.sub(r'[() ]', '', image1)
        if image1.endswith(".xlsx"):
            pass
        # Method 1: Direct assignment (most common)
        new_names[image1] = new_filename
        print(new_filename)

    # image_dir = sys.argv[1]
    saved_image_dir = r"C:\Users\Audrey\OneDrive\Pictures\screenshot-resized100"

    # i is the counter for number of dictionary items. The (image1, timestamp1) unpacks the dictionary items.
    try:
        for i, (image1, timestamp1) in enumerate(excel_fl.dictionary.items()):
            """
            uncomment this block of code if you want to debug or update the code, and only test with the first 10 images.
            """
            # if i >= 10:
            #     break
            """
            the os path join function is implemented so that every image listed in the dictionary automatically forms each filepath to each individual image
            in order to embed each thumbnail image into each row of the spreadsheet.
            """
            returned_path = excel_fl.filelist_thumbnail(idx=i, image=image1, image_dir=dir_path,
                                                        saved_image_dir=saved_image_dir)
            # print(returned_path)

            add_dict_to_spreadsheet(excel_fl.dictionary, "B", "C")
            # add_dict_to_spreadsheet(excel_fl.dictionary, "B", "D")
            # add_value_to_spreadsheet(new_names, "C")

            filelist_wb.save(fr'{dir_path}\{xlsx_name}_filelist.xlsx')
        print(f"\n{ws.title} spreadsheet is saved!")
    except:
        print("dictionary is empty!")
